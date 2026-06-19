from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import unicodedata
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader


CORPUS_ID = "awareness_knowledge_roadmap_security_2026_05_15"
POLICY_BOUNDARY = (
    "Use for knowledge, lab planning, detection engineering, and owned-scope "
    "preparation only. Live external testing, exploitation, credential access, "
    "destructive actions, third-party activity, and real-money trading require "
    "explicit scope, an applicable tool contract, approval, and a receipt."
)

CATEGORY_TERMS: dict[str, list[str]] = {
    "agentic_ai_security": [
        "agent",
        "agents",
        "agentic",
        "mcp",
        "llm",
        "gemini",
        "cursor",
        "hexstrike",
        "ai-driven",
        "prompt",
        "autonomous",
        "ollama",
        "openai",
        "claude",
        "langchain",
        "tool-calling",
        "orchestration",
        "hackerai",
    ],
    "security_lab_pentest": [
        "pentest",
        "penetration",
        "exploit",
        "burp",
        "nmap",
        "vulnerability",
        "red team",
        "idor",
        "xss",
        "csrf",
        "sql injection",
        "active directory",
        "kali",
        "metasploit",
        "bug bounty",
        "recon",
        "lateral movement",
        "privilege escalation",
        "payload",
    ],
    "detection_threat_intel": [
        "detection",
        "threat intelligence",
        "soc",
        "incident",
        "hunting",
        "insider",
        "apt",
        "ioc",
        "sigma",
        "yara",
        "ueba",
        "siem",
        "endpoint",
        "log source",
        "alert",
        "defender",
        "malicious",
        "telemetry",
        "correlation",
        "kill chain",
    ],
    "osint_attack_surface": [
        "osint",
        "shodan",
        "subdomain",
        "footprint",
        "attack surface",
        "exposure",
        "recon",
        "camera",
        "dork",
        "email",
        "leak",
        "asset discovery",
        "internet-exposed",
    ],
    "infrastructure_cloud_linux": [
        "kubernetes",
        "linux",
        "windows",
        "active directory",
        "cloud",
        "aws",
        "azure",
        "gcp",
        "docker",
        "ssh",
        "server",
        "kali",
        "network",
        "dns",
        "api",
        "websocket",
        "vm",
        "ubuntu",
        "endpoint",
    ],
    "career_roadmap_tools": [
        "roadmap",
        "course",
        "learn",
        "skills",
        "certification",
        "guide",
        "tools",
        "workflow",
        "career",
        "portfolio",
        "exam",
        "complete course",
        "tool stack",
    ],
}

CATEGORY_THRESHOLDS = {
    "agentic_ai_security": 2,
    "security_lab_pentest": 2,
    "detection_threat_intel": 2,
    "osint_attack_surface": 2,
    "infrastructure_cloud_linux": 2,
    "career_roadmap_tools": 3,
}

EXECUTION_TERMS = [
    "exploit",
    "exploitation",
    "payload",
    "credential",
    "password",
    "token",
    "api key",
    "nmap",
    "burp",
    "metasploit",
    "sqlmap",
    "shodan",
    "recon",
    "scan",
    "enumeration",
    "pentest",
    "penetration",
    "bug bounty",
    "idor",
    "xss",
    "csrf",
    "rce",
    "lateral movement",
    "privilege escalation",
    "active directory",
    "red team",
    "kali",
    "hexstrike",
    "malware",
]
SENSITIVE_TEXT_PATTERNS = [
    re.compile(r"sk-or-v1-[A-Za-z0-9_-]{20,}"),
    re.compile(r"eyJ[A-Za-z0-9_-]{20,}\.[A-Za-z0-9_-]{20,}\.[A-Za-z0-9_-]{20,}"),
]
REDACTION_MARKER = "[REDACTED_IMPORTED_SECRET_LIKE_VALUE]"

PLANES = {
    "agentic_ai_security": [
        "Tool / Automation Plane",
        "Knowledge / RAG Plane",
        "Understanding / Evaluation Plane",
    ],
    "security_lab_pentest": [
        "Security Lab Plane",
        "Knowledge / RAG Plane",
        "Understanding / Evaluation Plane",
    ],
    "detection_threat_intel": [
        "Security Lab Plane",
        "Knowledge / RAG Plane",
        "Understanding / Evaluation Plane",
    ],
    "osint_attack_surface": [
        "Security Lab Plane",
        "Knowledge / RAG Plane",
        "Understanding / Evaluation Plane",
    ],
    "infrastructure_cloud_linux": ["Security Lab Plane", "Knowledge / RAG Plane"],
    "career_roadmap_tools": ["Knowledge / RAG Plane", "Understanding / Evaluation Plane"],
    "general_awareness": ["Knowledge / RAG Plane"],
}

PLAYBOOKS = {
    "agentic_ai_security": [
        "PLAYBOOK_002_AGENT_MACHINE_WORKFLOW_INTELLIGENCE_INTAKE",
        "PLAYBOOK_004_TOOL_CONTRACT_RESOLVER",
    ],
    "security_lab_pentest": [
        "PLAYBOOK_005_SCOPE_AUTHORIZATION_PREPARATION",
        "PLAYBOOK_004_TOOL_CONTRACT_RESOLVER",
    ],
    "detection_threat_intel": [
        "PLAYBOOK_001_CORPUS_INTAKE_KNOWLEDGE_CONVERSION",
        "PLAYBOOK_003_RUNTIME_QUEUE_RECEIPT_LEDGER",
    ],
    "osint_attack_surface": [
        "PLAYBOOK_005_SCOPE_AUTHORIZATION_PREPARATION",
        "PLAYBOOK_004_TOOL_CONTRACT_RESOLVER",
    ],
    "infrastructure_cloud_linux": ["PLAYBOOK_005_SCOPE_AUTHORIZATION_PREPARATION"],
    "career_roadmap_tools": ["PLAYBOOK_001_CORPUS_INTAKE_KNOWLEDGE_CONVERSION"],
    "general_awareness": ["PLAYBOOK_001_CORPUS_INTAKE_KNOWLEDGE_CONVERSION"],
}


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def slugify(value: str, max_len: int = 72) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    ascii_value = re.sub(r"[^a-zA-Z0-9]+", "-", ascii_value.lower()).strip("-")
    if not ascii_value:
        ascii_value = "source"
    return ascii_value[:max_len].strip("-") or "source"


def rel(repo: Path, path: Path) -> str:
    return path.relative_to(repo).as_posix()


def safe_replace_dir(repo: Path, target: Path) -> None:
    target = target.resolve()
    repo = repo.resolve()
    if repo not in target.parents and target != repo:
        raise ValueError(f"Refusing to replace directory outside repo: {target}")
    if target.exists():
        shutil.rmtree(target)
    target.mkdir(parents=True, exist_ok=True)


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def redact_sensitive_text(text: str) -> tuple[str, int]:
    redactions = 0
    for pattern in SENSITIVE_TEXT_PATTERNS:
        text, count = pattern.subn(REDACTION_MARKER, text)
        redactions += count
    return text, redactions


def title_from_markdown(text: str, fallback: str) -> str:
    for line in text.splitlines():
        stripped = line.strip()
        if stripped.startswith("#"):
            return stripped.lstrip("#").strip()[:160] or fallback
    return fallback


def headings_from_markdown(text: str, max_count: int = 12) -> list[str]:
    headings: list[str] = []
    for line in text.splitlines():
        stripped = line.strip()
        if stripped.startswith("#"):
            heading = stripped.lstrip("#").strip()
            if heading:
                headings.append(heading[:220])
        if len(headings) >= max_count:
            break
    return headings


def published_from_text(text: str) -> str | None:
    match = re.search(
        r"Published\s+in\s+[^\n]+\s+on\s+([A-Z][a-z]+\s+\d{1,2},\s+\d{4})",
        text,
    )
    if not match:
        match = re.search(r"\b(20\d{2}-\d{2}-\d{2})\b", text)
        return match.group(1) if match else None
    for fmt in ("%b %d, %Y", "%B %d, %Y"):
        try:
            return datetime.strptime(match.group(1), fmt).date().isoformat()
        except ValueError:
            pass
    return match.group(1)


def term_score(text: str, terms: list[str]) -> int:
    lower = text.lower()
    score = 0
    for term in terms:
        escaped = re.escape(term.lower())
        if " " in term or "-" in term:
            score += len(re.findall(escaped, lower))
        else:
            score += len(re.findall(rf"\b{escaped}\b", lower))
    return score


def classify(title: str, headings: list[str], body_sample: str) -> list[str]:
    weighted = f"{title}\n" * 4 + ("\n".join(headings) + "\n") * 2 + body_sample[:30000]
    categories = []
    for category, terms in CATEGORY_TERMS.items():
        score = term_score(weighted, terms)
        if score >= CATEGORY_THRESHOLDS.get(category, 2):
            categories.append(category)
    return categories or ["general_awareness"]


def classify_sensitivity(title: str, headings: list[str], body_sample: str, categories: list[str]) -> str:
    weighted = f"{title}\n" * 3 + "\n".join(headings) + "\n" + body_sample[:30000]
    if term_score(weighted, EXECUTION_TERMS) >= 2:
        return "execution_capable"
    gated_categories = {"security_lab_pentest", "osint_attack_surface", "agentic_ai_security"}
    if gated_categories.intersection(categories):
        return "review_before_use"
    return "reference"


def planes_for(categories: list[str]) -> list[str]:
    ordered: list[str] = []
    for category in categories:
        for plane in PLANES.get(category, []):
            if plane not in ordered:
                ordered.append(plane)
    return ordered or ["Knowledge / RAG Plane"]


def playbooks_for(categories: list[str]) -> list[str]:
    ordered = ["PLAYBOOK_001_CORPUS_INTAKE_KNOWLEDGE_CONVERSION"]
    for category in categories:
        for playbook in PLAYBOOKS.get(category, []):
            if playbook not in ordered:
                ordered.append(playbook)
    if "PLAYBOOK_003_RUNTIME_QUEUE_RECEIPT_LEDGER" not in ordered:
        ordered.append("PLAYBOOK_003_RUNTIME_QUEUE_RECEIPT_LEDGER")
    return ordered


def bounded_dimension(value: int | None, default: int) -> int:
    if value is None or value < 1:
        return default
    return value


def convert_xlsx(path: Path, title: str) -> tuple[str, list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    lines = [
        f"# {title}",
        "",
        f"Converted from workbook `{path.name}` for searchable FATHIYA retrieval.",
        "",
    ]
    headings = [title]
    for ws in wb.worksheets[:8]:
        max_row = bounded_dimension(ws.max_row, 40)
        max_column = bounded_dimension(ws.max_column, 12)
        lines.append(f"## Sheet: {ws.title}")
        headings.append(f"Sheet: {ws.title}")
        lines.append(f"- Dimensions: {ws.max_row or 'unknown'} rows x {ws.max_column or 'unknown'} columns")
        sample_rows = []
        for row in ws.iter_rows(min_row=1, max_row=min(max_row, 40), values_only=True):
            values = [
                "" if value is None else str(value).replace("\n", " ")[:240]
                for value in row[: min(max_column, 12)]
            ]
            if any(values):
                sample_rows.append(values)
        if sample_rows:
            width = max(len(row) for row in sample_rows)
            lines.append("")
            lines.append("| " + " | ".join(f"c{i + 1}" for i in range(width)) + " |")
            lines.append("| " + " | ".join("---" for _ in range(width)) + " |")
            for row in sample_rows[:25]:
                padded = row + [""] * (width - len(row))
                escaped = [cell.replace("|", "\\|") for cell in padded]
                lines.append("| " + " | ".join(escaped) + " |")
        lines.append("")
    wb.close()
    return "\n".join(lines).strip() + "\n", headings[:12]


def convert_pdf(path: Path, title: str) -> tuple[str, list[str]]:
    reader = PdfReader(str(path))
    lines = [
        f"# {title}",
        "",
        f"Converted from PDF `{path.name}` for searchable FATHIYA retrieval.",
        "",
        f"- Pages: {len(reader.pages)}",
        "",
    ]
    chars = 0
    for idx, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        if not text.strip():
            continue
        lines.append(f"## Page {idx}")
        chunk = text.strip()[:8000]
        lines.append(chunk)
        lines.append("")
        chars += len(chunk)
        if chars >= 50000:
            lines.append("_PDF text truncated after 50,000 characters for compact retrieval._")
            break
    return "\n".join(lines).strip() + "\n", [title]


def build_entry(repo: Path, source_dir: Path, raw_root: Path, converted_root: Path, index: int, path: Path) -> dict[str, Any]:
    digest = sha256_file(path)
    ext = path.suffix.lower()
    source_id = f"aks-{index:03d}-{digest[:10]}"
    fallback_title = path.stem.replace("_", " ").replace("-", " ").strip()[:160]
    converted_path = None

    if ext == ".md":
        original_body = read_text(path)
        title = title_from_markdown(original_body, fallback_title)
        headings = headings_from_markdown(original_body)
        body, redaction_count = redact_sensitive_text(original_body)
        output = raw_root / f"{source_id}-{slugify(title)}.md"
        output.write_text(body, encoding="utf-8", newline="\n")
        body_sample = body
    elif ext == ".xlsx":
        title = fallback_title
        converted_body, headings = convert_xlsx(path, title)
        body, redaction_count = redact_sensitive_text(converted_body)
        output = converted_root / f"{source_id}-{slugify(title)}.md"
        output.write_text(body, encoding="utf-8", newline="\n")
        converted_path = rel(repo, output)
        body_sample = body
    elif ext == ".pdf":
        title = fallback_title
        converted_body, headings = convert_pdf(path, title)
        body, redaction_count = redact_sensitive_text(converted_body)
        output = converted_root / f"{source_id}-{slugify(title)}.md"
        output.write_text(body, encoding="utf-8", newline="\n")
        converted_path = rel(repo, output)
        body_sample = body
    else:
        raise ValueError(f"Unsupported extension: {path}")

    categories = classify(title, headings, body_sample)
    return {
        "source_id": source_id,
        "original_name": path.name,
        "extension": ext,
        "bytes": path.stat().st_size,
        "sha256": digest,
        "runtime_import_path": rel(repo, path),
        "knowledge_raw_path": rel(repo, output),
        "converted_path": converted_path,
        "title": title,
        "published": published_from_text(body_sample),
        "headings": headings,
        "categories": categories,
        "fathiya_planes": planes_for(categories),
        "sensitivity": classify_sensitivity(title, headings, body_sample, categories),
        "content_redactions": redaction_count,
        "policy_boundary": POLICY_BOUNDARY,
        "recommended_playbooks": playbooks_for(categories),
    }


def write_manifest(repo: Path, zip_path: Path | None, source_dir: Path, raw_root: Path, entries: list[dict[str, Any]]) -> Path:
    manifest_path = repo / "knowledge" / "intake" / "runtime" / f"{CORPUS_ID}_manifest.json"
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    category_counts = Counter(category for entry in entries for category in entry["categories"])
    plane_counts = Counter(plane for entry in entries for plane in entry["fathiya_planes"])
    sensitivity_counts = Counter(entry["sensitivity"] for entry in entries)
    extension_counts = Counter(entry["extension"] for entry in entries)
    redaction_count = sum(int(entry.get("content_redactions") or 0) for entry in entries)
    source_zip_hash = sha256_file(zip_path) if zip_path and zip_path.exists() else None
    manifest = {
        "schema_version": "FATHIYA_IMPORTED_CORPUS_MANIFEST_v1",
        "corpus_id": CORPUS_ID,
        "source_zip": str(zip_path) if zip_path else None,
        "source_zip_sha256": source_zip_hash,
        "imported_at": datetime.now(timezone.utc).isoformat(),
        "raw_runtime_root": rel(repo, source_dir),
        "knowledge_raw_root": rel(repo, raw_root),
        "policy_boundary": POLICY_BOUNDARY,
        "totals": {
            "files": len(entries),
            "extensions": dict(extension_counts),
            "categories": dict(category_counts.most_common()),
            "planes": dict(plane_counts.most_common()),
            "sensitivity": dict(sensitivity_counts.most_common()),
            "content_redactions": redaction_count,
        },
        "entries": entries,
    }
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    return manifest_path


def write_registry(repo: Path, manifest_path: Path, raw_root: Path, entries: list[dict[str, Any]], zip_path: Path | None) -> Path:
    registry_path = repo / "knowledge" / "registries" / "imported_corpus_registry_v1.json"
    registry_path.parent.mkdir(parents=True, exist_ok=True)
    category_counts = Counter(category for entry in entries for category in entry["categories"])
    plane_counts = Counter(plane for entry in entries for plane in entry["fathiya_planes"])
    source_zip_hash = sha256_file(zip_path) if zip_path and zip_path.exists() else None
    registry = {
        "schema_version": "FATHIYA_IMPORTED_CORPUS_REGISTRY_v1",
        "status": "active",
        "purpose": "Register operator-provided external corpora that should inform FATHIYA agents through retrieval and policy-gated playbooks.",
        "corpora": [
            {
                "corpus_id": CORPUS_ID,
                "title": "Awareness, Knowledge Roadmap, and Security Corpus",
                "source_zip_sha256": source_zip_hash,
                "manifest_path": rel(repo, manifest_path),
                "knowledge_raw_root": rel(repo, raw_root),
                "file_count": len(entries),
                "dominant_categories": list(category_counts.keys()),
                "fathiya_planes": list(plane_counts.keys()),
                "policy_boundary": POLICY_BOUNDARY,
                "default_mode": "Knowledge + Lab/Internal Owned Surface Preparation",
                "approval_required_for": [
                    "live external scanning or probing",
                    "exploitation or vulnerability validation",
                    "credential access or secret handling",
                    "destructive action",
                    "third-party target activity",
                    "external webhooks, workflow activation, or publication",
                    "real-money trading or broker order placement",
                ],
            }
        ],
    }
    registry_path.write_text(
        json.dumps(registry, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
        newline="\n",
    )
    return registry_path


def write_report(repo: Path, manifest_path: Path, registry_path: Path, raw_root: Path, entries: list[dict[str, Any]]) -> Path:
    report_path = repo / "knowledge" / "reports" / "study" / "FATHIYA_AWARENESS_SECURITY_CORPUS_INGEST_REPORT_v1.md"
    report_path.parent.mkdir(parents=True, exist_ok=True)
    category_counts = Counter(category for entry in entries for category in entry["categories"])
    plane_counts = Counter(plane for entry in entries for plane in entry["fathiya_planes"])
    sensitivity_counts = Counter(entry["sensitivity"] for entry in entries)
    extension_counts = Counter(entry["extension"] for entry in entries)
    redaction_count = sum(int(entry.get("content_redactions") or 0) for entry in entries)
    largest = sorted(entries, key=lambda entry: entry["bytes"], reverse=True)[:12]
    converted_root = raw_root / "converted"

    lines = [
        "# FATHIYA Awareness/Security Corpus Ingest Report v1",
        "",
        "## Verdict",
        "",
        "The uploaded `AWARENESS_KNOWLEDGE_ROADMAP-and-securty.zip` is now a primary imported operator corpus for FATHIYA. Markdown sources were copied into `knowledge/raw/imports/...` using stable ASCII filenames, while every original filename is preserved in the manifest. Spreadsheet and PDF files were converted into searchable Markdown summaries. Raw extracted copies remain in the local runtime import directory.",
        "",
        "This corpus contains execution-capable security material. It is for learning, lab planning, owned-scope preparation, detection engineering, and agent/tool design. Live testing, exploitation, credential access, destructive actions, third-party activity, and real-money trading remain blocked unless FATHIYA has target scope, an applicable tool contract, explicit approval, and a receipt.",
        "",
        "## Inventory",
        "",
        f"- Source files: {len(entries)}",
    ]
    lines.extend(f"- `{ext}`: {count}" for ext, count in extension_counts.most_common())
    lines.extend(["", "## Dominant Categories", ""])
    lines.extend(f"- `{category}`: {count}" for category, count in category_counts.most_common())
    lines.extend(["", "## FATHIYA Planes Affected", ""])
    lines.extend(f"- {plane}: {count}" for plane, count in plane_counts.most_common())
    lines.extend(["", "## Sensitivity", ""])
    lines.extend(f"- `{name}`: {count}" for name, count in sensitivity_counts.most_common())
    lines.extend(
        [
            "",
            "## Content Redactions",
            "",
            f"- Secret-like imported text values redacted from searchable copies: {redaction_count}",
            "- Original file names and source SHA-256 hashes remain preserved in the manifest.",
        ]
    )
    lines.extend(["", "## Largest / Highest-Weight Sources", ""])
    lines.extend(
        f"- `{entry['source_id']}` - {entry['title']} (`{entry['extension']}`, {entry['bytes']} bytes) -> `{entry['knowledge_raw_path']}`"
        for entry in largest
    )
    lines.extend(
        [
            "",
            "## Operating Interpretation",
            "",
            "- AI-agent, MCP, Cursor, Gemini, HexStrike, prompt-injection, and autonomous-security material routes through PLAYBOOK_002 and PLAYBOOK_004 before it becomes a tool, workflow, or agent capability.",
            "- Pentest, recon, OSINT, AD, cloud, Kubernetes, Burp, Nmap, IDOR, upload, and bug-bounty material routes through PLAYBOOK_005 and remains preparation-only unless a Target Card and written policy authorize activity.",
            "- Detection engineering, threat intelligence, SOC, endpoint, and incident material can be used immediately for defensive lab plans, detections, report templates, and knowledge cards.",
            "- Career, roadmap, tool-stack, and learning material feeds the Knowledge/RAG and Understanding/Evaluation planes.",
            "- Execution-capable snippets are not execution permission. They become lab notes, checklists, local-only examples, or blocked approval items depending on scope.",
            "",
            "## Generated Artifacts",
            "",
            f"- Manifest: `{rel(repo, manifest_path)}`",
            f"- Registry: `{rel(repo, registry_path)}`",
            f"- Raw Markdown import root: `{rel(repo, raw_root)}`",
            f"- Converted spreadsheet/PDF summaries: `{rel(repo, converted_root)}`",
            "",
            "## Comprehension Check Prompt",
            "",
            "```text",
            "استرجع corpus awareness_knowledge_roadmap_security_2026_05_15 واصنع خريطة فهم: AI agents، security lab، OSINT/recon، detection engineering، tool contracts، وما هو مسموح أو ممنوع قبل التنفيذ. سجل إيصالًا.",
            "```",
            "",
        ]
    )
    report_path.write_text("\n".join(lines), encoding="utf-8", newline="\n")
    return report_path


def import_corpus(repo: Path, source_dir: Path, zip_path: Path | None) -> dict[str, Any]:
    raw_root = repo / "knowledge" / "raw" / "imports" / CORPUS_ID
    converted_root = raw_root / "converted"
    safe_replace_dir(repo, raw_root)
    converted_root.mkdir(parents=True, exist_ok=True)

    files = sorted(
        [path for path in source_dir.iterdir() if path.is_file() and path.suffix.lower() in {".md", ".xlsx", ".pdf"}],
        key=lambda path: path.name.casefold(),
    )
    entries = [
        build_entry(repo, source_dir, raw_root, converted_root, index, path)
        for index, path in enumerate(files, start=1)
    ]
    manifest_path = write_manifest(repo, zip_path, source_dir, raw_root, entries)
    registry_path = write_registry(repo, manifest_path, raw_root, entries, zip_path)
    report_path = write_report(repo, manifest_path, registry_path, raw_root, entries)
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    return {
        "corpus_id": CORPUS_ID,
        "entries": len(entries),
        "manifest": rel(repo, manifest_path),
        "registry": rel(repo, registry_path),
        "report": rel(repo, report_path),
        "raw_root": rel(repo, raw_root),
        "totals": manifest["totals"],
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import the FATHIYA awareness/security ZIP corpus into knowledge.")
    parser.add_argument("--repo", type=Path, default=Path.cwd())
    parser.add_argument("--source-dir", type=Path, required=True)
    parser.add_argument("--zip-path", type=Path)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    repo = args.repo.resolve()
    source_dir = args.source_dir.resolve()
    if not source_dir.exists():
        raise SystemExit(f"Source directory does not exist: {source_dir}")
    zip_path = args.zip_path.resolve() if args.zip_path else None
    result = import_corpus(repo, source_dir, zip_path)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
